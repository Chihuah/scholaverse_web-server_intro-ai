"""Excel import service — parses TronClass Excel reports into structured records.

Two report types are supported:
- Completion report (完成度_xxx.xlsx): completion_rate only
- Score list report (score_list.xlsx): pretest_score + quiz_score
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from typing import IO

import openpyxl

logger = logging.getLogger(__name__)

# ─── Constants ────────────────────────────────────────────────────────

COMPLETION_HEADER_MAP: dict[str, str] = {
    "1-人工智慧與深度學習的基礎": "unit_1",
    "2-多層感知器 - 回歸與分類問題": "unit_2",
    "3-卷積神經網路 - 電腦視覺": "unit_3",
    "4-循環神經網路 - 自然語言處理": "unit_4",
    "5-建構深度學習網路模型": "unit_5",
    # unit_6（自主學習）的完成度不採用 TronClass 課程完成度，
    # 改由 score_list 作業成績加權平均計算（見 HOMEWORK_ASSIGNMENTS），
    # 因此完成度報表的「6-自主學習」欄位刻意不匯入。
}

# unit_6（自主學習）完成度 = 全部作業加權平均。
# 作業名稱須與 score_list.xlsx 第 2 列標頭一致（去除 "(x.xx%)" 後綴後比對）。
# 權重採課程總成績原始占比：第一個作業 6%，其餘各 7%，合計 55%。
HOMEWORK_ASSIGNMENTS: dict[str, float] = {
    "初玩類神經網路": 6.0,
    "實作單元感知器: NAND與NOR邏輯閘": 7.0,
    "MLP預測波士頓房價": 7.0,
    "多層感知器實作案例: 鐵達尼號乘客資料集": 7.0,
    "卷積神經網路辨識MNIST手寫數字資料集": 7.0,
    "卷積神經網路辨識cifar-10資料集": 7.0,
    "循環神經網路實作: 特斯拉公司美股股價預測": 7.0,
    "自主學習: AI 工具實際操作應用成果報告": 7.0,
}

HOMEWORK_TOTAL_WEIGHT: float = sum(HOMEWORK_ASSIGNMENTS.values())

CHAPTER_NUMBER_MAP: dict[str, str] = {
    "第一章": "unit_1",
    "第二章": "unit_2",
    "第三章": "unit_3",
    "第四章": "unit_4",
    "第五章": "unit_5",
    "第六章": "unit_6",
}

# ─── Data Structures ──────────────────────────────────────────────────


@dataclass
class StudentRecord:
    student_id: str
    unit_code: str
    completion_rate: float | None = None
    quiz_score: float | None = None
    preview_score: float | None = None
    pretest_score: float | None = None


@dataclass
class HomeworkRecord:
    student_id: str
    assignment_name: str
    score: float


@dataclass
class ExcelParseResult:
    records: list[StudentRecord] = field(default_factory=list)
    homework: list[HomeworkRecord] = field(default_factory=list)
    unrecognized_headers: list[str] = field(default_factory=list)
    parse_errors: list[str] = field(default_factory=list)


# ─── Value Parsers ────────────────────────────────────────────────────


def _parse_completion_rate(val: object) -> float | None:
    """'87.5%' → 87.5,  '—'/None → None."""
    if val is None:
        return None
    s = str(val).strip()
    if s in ("—", ""):
        return None
    if s.endswith("%"):
        try:
            return float(s[:-1])
        except ValueError:
            return None
    return None


def _parse_quiz_completion(val: object) -> float | None:
    """'100.0分' → 100.0,  '未完成' → 0.0,  '—'/None → None."""
    if val is None:
        return None
    s = str(val).strip()
    if s in ("—", ""):
        return None
    if s == "未完成":
        return 0.0
    if s.endswith("分"):
        try:
            return float(s[:-1])
        except ValueError:
            return None
    return None


def _parse_score(val: object) -> float | None:
    """Numeric string → float,  '未繳'/'未批改'/'—'/None → None."""
    if val is None:
        return None
    s = str(val).strip()
    if s in ("未繳", "未批改", "—", ""):
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _parse_homework_score(val: object) -> float | None:
    """作業分數解析：未繳/未批改一律以 0 分計（拉低平均），空欄/無法解析 → None。"""
    if val is None:
        return None
    s = str(val).strip()
    if s in ("未繳", "未批改"):
        return 0.0
    if s in ("—", ""):
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def compute_homework_completion(scores_by_assignment: dict[str, float | None]) -> float:
    """unit_6 完成度 = Σ(作業分數 × 權重) ÷ 總權重，缺漏作業以 0 分計。

    分母固定為 HOMEWORK_TOTAL_WEIGHT（含尚未上線的第六章作業），
    符合「未繳給 0 分拉低平均」的公平原則。
    """
    earned = sum(
        (scores_by_assignment.get(name) or 0.0) * weight
        for name, weight in HOMEWORK_ASSIGNMENTS.items()
    )
    return round(earned / HOMEWORK_TOTAL_WEIGHT, 1)


# ─── Chapter Resolver ─────────────────────────────────────────────────


def _chapter_unit_code(header: str) -> str | None:
    """Return unit_code from a header containing a chapter number like '第一章'."""
    for chapter, unit_code in CHAPTER_NUMBER_MAP.items():
        if chapter in header:
            return unit_code
    return None


# ─── Public API ───────────────────────────────────────────────────────


def parse_completion_excel(file_bytes: bytes) -> ExcelParseResult:
    """Parse a TronClass completion-rate Excel report.

    Expected layout:
    - Row 1: headers (single-row header)
    - Col 2 (index 1): student account / student_id
    - Data starts at row 2
    """
    result = ExcelParseResult()

    try:
        wb = openpyxl.load_workbook(filename=__import__("io").BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        result.parse_errors.append(f"無法讀取 Excel 檔案：{exc}")
        return result

    ws = wb.active

    # Read header row (row 1)
    headers = [cell.value for cell in ws[1]]

    # Column mappings: list of (col_idx, unit_code, field_name)
    col_mappings: list[tuple[int, str, str]] = []

    # Note: quiz score columns ("第X章 課後測驗") in this report only carry a
    # binary completion marker ("100.0分" / "未完成"), not the real score.
    # We deliberately ignore them here — actual quiz scores come from
    # score_list.xlsx via parse_score_excel().
    for idx, header in enumerate(headers):
        if header is None:
            continue
        h = str(header).strip()

        # Completion rate column
        if h in COMPLETION_HEADER_MAP:
            col_mappings.append((idx, COMPLETION_HEADER_MAP[h], "completion_rate"))
            continue

        # Quietly skip the post-test marker columns (handled by score_list).
        if "課後測驗" in h:
            continue

        # Skip known non-data columns (rank, name, student id col, etc.)
        if idx <= 1:
            continue

        result.unrecognized_headers.append(h)

    # Process data rows starting at row 2
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Skip fully empty rows
        if all(v is None for v in row):
            continue

        student_id = row[1] if len(row) > 1 else None
        if student_id is None or str(student_id).strip() == "":
            continue
        student_id = str(student_id).strip()

        for col_idx, unit_code, _field_name in col_mappings:
            raw = row[col_idx] if col_idx < len(row) else None
            parsed = _parse_completion_rate(raw)
            if parsed is None:
                continue

            existing = _find_record(result.records, student_id, unit_code)
            if existing is None:
                existing = StudentRecord(student_id=student_id, unit_code=unit_code)
                result.records.append(existing)
            existing.completion_rate = parsed

    return result


def parse_score_excel(file_bytes: bytes) -> ExcelParseResult:
    """Parse a TronClass score-list Excel report.

    Expected layout:
    - Row 1: group headers (ignored)
    - Row 2: column headers (double-row header)
    - Col 1 (index 0): student account / student_id
    - Data starts at row 3
    """
    result = ExcelParseResult()

    try:
        wb = openpyxl.load_workbook(filename=__import__("io").BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        result.parse_errors.append(f"無法讀取 Excel 檔案：{exc}")
        return result

    ws = wb.active

    # Read header row 2 (index 1)
    header_row = ws[2]
    headers = [cell.value for cell in header_row]

    col_mappings: list[tuple[int, str, str]] = []

    for idx, header in enumerate(headers):
        if header is None:
            continue
        h = str(header).strip()

        # Strip percentage suffix like "(40%)" or "(6.00%)" from header names
        h_clean = re.sub(r'\(\d+(?:\.\d+)?%\)', '', h).strip()

        # Homework assignment columns (drive unit_6 completion)
        if h_clean in HOMEWORK_ASSIGNMENTS:
            col_mappings.append((idx, h_clean, "homework"))
            continue

        if "前測" in h_clean:
            unit_code = _chapter_unit_code(h_clean)
            if unit_code:
                col_mappings.append((idx, unit_code, "pretest_score"))
                continue

        if "課後測驗" in h_clean:
            unit_code = _chapter_unit_code(h_clean)
            if unit_code:
                col_mappings.append((idx, unit_code, "quiz_score"))
                continue

        # Skip known non-data columns (rank, name, student id col, etc.)
        if idx <= 0:
            continue

        result.unrecognized_headers.append(h)

    # Process data rows starting at row 3
    for row in ws.iter_rows(min_row=3, values_only=True):
        # Skip fully empty rows
        if all(v is None for v in row):
            continue

        student_id = row[0] if len(row) > 0 else None
        if student_id is None or str(student_id).strip() == "":
            continue

        student_id_str = str(student_id).strip()

        # Skip summary / footer rows (non-numeric student IDs that look like labels)
        # A valid student_id should be purely numeric or match a known pattern
        # We skip rows where student_id contains Chinese characters
        if re.search(r'[\u4e00-\u9fff]', student_id_str):
            continue

        for col_idx, key, field_name in col_mappings:
            raw = row[col_idx] if col_idx < len(row) else None

            if field_name == "homework":
                hw_score = _parse_homework_score(raw)
                if hw_score is None:
                    continue
                result.homework.append(
                    HomeworkRecord(
                        student_id=student_id_str,
                        assignment_name=key,
                        score=hw_score,
                    )
                )
                continue

            parsed = _parse_score(raw)
            if parsed is None:
                continue

            existing = _find_record(result.records, student_id_str, key)
            if existing is None:
                existing = StudentRecord(student_id=student_id_str, unit_code=key)
                result.records.append(existing)

            if field_name == "pretest_score":
                existing.pretest_score = parsed
            else:
                existing.quiz_score = parsed

    return result


# ─── Internal Helpers ─────────────────────────────────────────────────


def _find_record(
    records: list[StudentRecord], student_id: str, unit_code: str
) -> StudentRecord | None:
    for r in records:
        if r.student_id == student_id and r.unit_code == unit_code:
            return r
    return None

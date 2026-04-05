#!/usr/bin/env python3
"""Export per-student preview rates from TronClass completion snapshots.

Usage:
    uv run python scripts/export_preview_rates.py

The script reads files from the same directory as this script:
    - checkpoint.xlsx
    - 12-digit snapshot files such as 202603190700.xlsx

It writes:
    - preview_rates.csv
"""

from __future__ import annotations

import csv
import re
import sys
from dataclasses import dataclass, replace
from datetime import date, datetime, time, timedelta
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SNAPSHOT_FILENAME_RE = re.compile(r"^(?P<timestamp>\d{12})\.xlsx$")
VIDEO_CODE_RE = re.compile(r"^(?P<chapter>\d+)-(?P<section>\d+)-(?P<video>\d+)\b")
PROGRESS_RE = re.compile(r"(?P<value>\d+(?:\.\d+)?)%")
CHECKPOINT_FILENAME = "checkpoint.xlsx"
OUTPUT_FILENAME = "preview_rates.csv"
ACCOUNT_HEADERS = ("帳號", "學號")
PASSING_PROGRESS = 0.8
DATETIME_FORMAT = "%Y-%m-%d %H:%M"
HELP_TEXT = """\
影片預習率匯出腳本

用途:
  從專案 data/ 目錄中的 checkpoint.xlsx 與多個完成度快照 .xlsx
  計算每位學生、每單元的影片預習率，輸出為 preview_rates.csv。

執行方式:
  cd /var/www/app.scholaverse.cc/intro-ai
  .venv/bin/python scripts/export_preview_rates.py

可選說明:
  .venv/bin/python scripts/export_preview_rates.py --help

檔案放置位置:
  /var/www/app.scholaverse.cc/intro-ai/data/checkpoint.xlsx
  /var/www/app.scholaverse.cc/intro-ai/data/YYYYMMDDHHMM.xlsx

注意事項:
  - 只會讀取 data/ 目錄下的 .xlsx 檔案。
  - scholaverse.db 會被忽略，其他非 .xlsx 檔案也會被忽略。
  - checkpoint.xlsx 必須包含 video 與 time 欄位。
  - 完成度快照檔名必須是 12 碼時間戳，例如 202603190700.xlsx。
  - 結果會輸出到 /var/www/app.scholaverse.cc/intro-ai/data/preview_rates.csv
"""


class PreviewRateError(RuntimeError):
    """Raised when the input data is invalid or incomplete."""


@dataclass(frozen=True)
class CheckpointVideo:
    """Checkpoint metadata for a single video."""

    code: str
    title: str
    checkpoint_at: datetime
    unit_code: str


@dataclass(frozen=True)
class CompletionSnapshot:
    """Parsed completion snapshot workbook."""

    path: Path
    snapshot_at: datetime
    account_header: str
    video_columns: dict[str, int]
    student_rows: dict[str, list[Any]]


@dataclass(frozen=True)
class PreviewRateRow:
    """Output CSV row."""

    student_id: str
    unit_code: str
    preview_score: str
    eligible_video_count: int
    previewed_video_count: int
    latest_checkpoint_at: str
    source_snapshot_at: str


def get_workspace_dir() -> Path:
    """Return the project data directory used by this script."""

    return Path(__file__).resolve().parent.parent / "data"


def excel_date_to_datetime(value: Any) -> datetime:
    """Convert Excel cell values to a datetime."""

    if isinstance(value, datetime):
        return value.replace(second=0, microsecond=0)

    if isinstance(value, date):
        return datetime.combine(value, time.min)

    if isinstance(value, (int, float)):
        base = datetime(1899, 12, 30)
        converted = base + timedelta(days=float(value))
        return converted.replace(second=0, microsecond=0)

    if isinstance(value, str):
        text = value.strip()
        if not text:
            raise PreviewRateError("Checkpoint time cannot be blank when parsed.")
        for fmt in ("%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M", "%Y-%m-%d", "%Y/%m/%d"):
            try:
                parsed = datetime.strptime(text, fmt)
                return parsed.replace(second=0, microsecond=0)
            except ValueError:
                continue
        raise PreviewRateError(f"Unsupported checkpoint time value: {value!r}")

    raise PreviewRateError(f"Unsupported checkpoint time value type: {type(value)!r}")


def extract_video_code(value: Any) -> str | None:
    """Extract a video code such as 1-2-01 from a cell value."""

    if value is None:
        return None

    text = str(value).strip()
    if not text:
        return None

    match = VIDEO_CODE_RE.match(text)
    if not match:
        return None

    return match.group(0)


def unit_code_from_video_code(video_code: str) -> str:
    """Map a video code to a unit code."""

    chapter = video_code.split("-", maxsplit=1)[0]
    return f"unit_{chapter}"


def parse_progress(value: Any) -> float:
    """Convert workbook progress cells into a numeric completion ratio."""

    if value is None:
        return 0.0

    text = str(value).strip()
    if not text or text in {"—", "未完成"}:
        return 0.0
    if text == "已完成":
        return 1.0

    match = PROGRESS_RE.search(text)
    if match:
        return float(match.group("value")) / 100.0

    raise PreviewRateError(f"Unsupported progress value: {value!r}")


def read_first_sheet_rows(path: Path) -> list[list[Any]]:
    """Read the first worksheet into a list of rows."""

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        return [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def build_header_index(header_row: list[Any], path: Path) -> dict[str, int]:
    """Build a lookup for non-empty header cells."""

    index: dict[str, int] = {}
    for column_index, raw_header in enumerate(header_row):
        if raw_header is None:
            continue
        header = str(raw_header).strip()
        if not header:
            continue
        index.setdefault(header, column_index)
    return index


def load_checkpoint_rows(path: Path) -> dict[str, CheckpointVideo]:
    """Load checkpoint videos from checkpoint.xlsx."""

    rows = read_first_sheet_rows(path)
    if not rows:
        raise PreviewRateError(f"{path.name} is empty.")

    header_index = build_header_index(rows[0], path)
    missing_headers = [name for name in ("video", "time") if name not in header_index]
    if missing_headers:
        raise PreviewRateError(
            f"{path.name} is missing required columns: {', '.join(missing_headers)}."
        )

    checkpoint_videos: dict[str, CheckpointVideo] = {}
    for row_number, row in enumerate(rows[1:], start=2):
        raw_title = row[header_index["video"]] if header_index["video"] < len(row) else None
        raw_time = row[header_index["time"]] if header_index["time"] < len(row) else None

        if raw_title is None or not str(raw_title).strip():
            continue

        if raw_time is None or str(raw_time).strip() == "":
            continue

        video_code = extract_video_code(raw_title)
        if not video_code:
            raise PreviewRateError(
                f"{path.name}:{row_number} has a video title without a code: {raw_title!r}."
            )

        if video_code in checkpoint_videos:
            raise PreviewRateError(
                f"{path.name}:{row_number} duplicates video code {video_code}."
            )

        checkpoint_videos[video_code] = CheckpointVideo(
            code=video_code,
            title=str(raw_title).strip(),
            checkpoint_at=excel_date_to_datetime(raw_time),
            unit_code=unit_code_from_video_code(video_code),
        )

    return checkpoint_videos


def pick_account_header(header_index: dict[str, int], path: Path) -> str:
    """Select the preferred account column, allowing backward-compatible fallback."""

    for header in ACCOUNT_HEADERS:
        if header in header_index:
            if header != ACCOUNT_HEADERS[0]:
                print(
                    f"[warn] {path.name} uses {header!r} instead of "
                    f"{ACCOUNT_HEADERS[0]!r}; treating it as student_id."
                )
            return header

    raise PreviewRateError(
        f"{path.name} is missing the required account column {ACCOUNT_HEADERS[0]!r}."
    )


def load_completion_snapshot(path: Path) -> CompletionSnapshot:
    """Load a completion snapshot workbook."""

    match = SNAPSHOT_FILENAME_RE.match(path.name)
    if not match:
        raise PreviewRateError(f"Invalid snapshot filename: {path.name}")

    snapshot_at = datetime.strptime(match.group("timestamp"), "%Y%m%d%H%M")
    rows = read_first_sheet_rows(path)
    if not rows:
        raise PreviewRateError(f"{path.name} is empty.")

    header_index = build_header_index(rows[0], path)
    account_header = pick_account_header(header_index, path)

    video_columns: dict[str, int] = {}
    for header, column_index in header_index.items():
        video_code = extract_video_code(header)
        if not video_code:
            continue
        if video_code in video_columns:
            raise PreviewRateError(f"{path.name} contains duplicate video code {video_code}.")
        video_columns[video_code] = column_index

    student_rows: dict[str, list[Any]] = {}
    account_column = header_index[account_header]
    for row_number, row in enumerate(rows[1:], start=2):
        if account_column >= len(row):
            continue
        raw_student_id = row[account_column]
        if raw_student_id is None:
            continue
        student_id = str(raw_student_id).strip()
        if not student_id:
            continue
        if student_id in student_rows:
            raise PreviewRateError(f"{path.name}:{row_number} duplicates account {student_id!r}.")
        student_rows[student_id] = row

    return CompletionSnapshot(
        path=path,
        snapshot_at=snapshot_at,
        account_header=account_header,
        video_columns=video_columns,
        student_rows=student_rows,
    )


def canonicalize_student_id(student_id: str) -> str:
    """Return a comparable student key."""

    text = student_id.strip()
    if text.isdigit() and len(text) > 4:
        return text[-4:]
    return text


def align_student_sets(snapshots: list[CompletionSnapshot]) -> tuple[list[CompletionSnapshot], list[str]]:
    """Align snapshots that mix full account IDs and short student-number suffixes."""

    if not snapshots:
        return [], []

    raw_to_canonical: dict[str, str] = {}
    canonical_to_output: dict[str, str] = {}

    aligned_snapshots: list[CompletionSnapshot] = []
    for snapshot in snapshots:
        remapped_rows: dict[str, list[Any]] = {}
        for raw_student_id, row in snapshot.student_rows.items():
            canonical_id = canonicalize_student_id(raw_student_id)
            if canonical_id in remapped_rows:
                raise PreviewRateError(
                    f"{snapshot.path.name} maps multiple rows to the same canonical student id "
                    f"{canonical_id!r}."
                )
            remapped_rows[canonical_id] = row
            raw_to_canonical[raw_student_id] = canonical_id

            preferred_output = canonical_to_output.get(canonical_id)
            if preferred_output is None or len(raw_student_id) > len(preferred_output):
                canonical_to_output[canonical_id] = raw_student_id

        aligned_snapshots.append(replace(snapshot, student_rows=remapped_rows))

    baseline = set(aligned_snapshots[0].student_rows)
    for snapshot in aligned_snapshots[1:]:
        current = set(snapshot.student_rows)
        if current != baseline:
            missing = sorted(baseline - current)
            extra = sorted(current - baseline)
            raise PreviewRateError(
                f"{snapshot.path.name} has mismatched students. "
                f"Missing: {missing or '[]'}; Extra: {extra or '[]'}."
            )

    ordered_student_ids = [canonical_to_output[student_id] for student_id in sorted(baseline)]
    return aligned_snapshots, ordered_student_ids


def select_snapshot_for_checkpoint(
    video_code: str, checkpoint_at: datetime, snapshots: list[CompletionSnapshot]
) -> CompletionSnapshot:
    """Select the latest snapshot at or before the checkpoint that contains the video."""

    candidates = [
        snapshot
        for snapshot in snapshots
        if snapshot.snapshot_at <= checkpoint_at and video_code in snapshot.video_columns
    ]
    if not candidates:
        raise PreviewRateError(
            f"No snapshot found for video {video_code} at or before "
            f"{checkpoint_at.strftime(DATETIME_FORMAT)}."
        )
    return max(candidates, key=lambda snapshot: snapshot.snapshot_at)


def round_score(value: float) -> str:
    """Round a score to two decimal places using decimal half-up."""

    return str(
        Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    )


def format_dt(value: datetime) -> str:
    """Format datetimes consistently."""

    return value.strftime(DATETIME_FORMAT)


def build_student_unit_rows(
    checkpoint_videos: dict[str, CheckpointVideo],
    snapshots: list[CompletionSnapshot],
) -> list[PreviewRateRow]:
    """Build output rows grouped by student and unit."""

    aligned_snapshots, student_ids = align_student_sets(snapshots)
    selected_snapshots: dict[str, CompletionSnapshot] = {}
    for video_code, checkpoint in checkpoint_videos.items():
        selected_snapshots[video_code] = select_snapshot_for_checkpoint(
            video_code, checkpoint.checkpoint_at, aligned_snapshots
        )

    unit_to_videos: dict[str, list[CheckpointVideo]] = {}
    for checkpoint in checkpoint_videos.values():
        unit_to_videos.setdefault(checkpoint.unit_code, []).append(checkpoint)

    for unit_code, videos in sorted(unit_to_videos.items()):
        snapshot_times = {
            selected_snapshots[video.code].snapshot_at.strftime(DATETIME_FORMAT)
            for video in videos
        }
        if len(snapshot_times) > 1:
            joined = ", ".join(sorted(snapshot_times))
            print(f"[info] {unit_code} uses multiple source snapshots: {joined}")

    output_rows: list[PreviewRateRow] = []
    for student_id in student_ids:
        canonical_student_id = canonicalize_student_id(student_id)
        for unit_code in sorted(unit_to_videos):
            videos = sorted(unit_to_videos[unit_code], key=lambda item: item.code)
            if not videos:
                continue

            latest_checkpoint_at = max(video.checkpoint_at for video in videos)
            source_snapshot_at = max(
                selected_snapshots[video.code].snapshot_at for video in videos
            )
            previewed_count = 0
            for video in videos:
                snapshot = selected_snapshots[video.code]
                row = snapshot.student_rows[canonical_student_id]
                column_index = snapshot.video_columns[video.code]
                raw_value = row[column_index] if column_index < len(row) else None
                if parse_progress(raw_value) >= PASSING_PROGRESS:
                    previewed_count += 1

            eligible_count = len(videos)
            preview_score = (previewed_count / eligible_count) * 100 if eligible_count else 0.0

            output_rows.append(
                PreviewRateRow(
                    student_id=student_id,
                    unit_code=unit_code,
                    preview_score=round_score(preview_score),
                    eligible_video_count=eligible_count,
                    previewed_video_count=previewed_count,
                    latest_checkpoint_at=format_dt(latest_checkpoint_at),
                    source_snapshot_at=format_dt(source_snapshot_at),
                )
            )

    output_rows.sort(key=lambda item: (item.student_id, item.unit_code))
    return output_rows


def write_csv(path: Path, rows: list[PreviewRateRow]) -> None:
    """Write preview rate rows as a UTF-8 BOM CSV."""

    with path.open("w", newline="", encoding="utf-8-sig") as csv_file:
        writer = csv.DictWriter(
            csv_file,
            fieldnames=[
                "student_id",
                "unit_code",
                "preview_score",
                "eligible_video_count",
                "previewed_video_count",
                "latest_checkpoint_at",
                "source_snapshot_at",
            ],
        )
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    "student_id": row.student_id,
                    "unit_code": row.unit_code,
                    "preview_score": row.preview_score,
                    "eligible_video_count": row.eligible_video_count,
                    "previewed_video_count": row.previewed_video_count,
                    "latest_checkpoint_at": row.latest_checkpoint_at,
                    "source_snapshot_at": row.source_snapshot_at,
                }
            )


def collect_snapshot_paths(workspace_dir: Path) -> list[Path]:
    """Collect snapshot workbooks from the workspace directory."""

    paths = [path for path in workspace_dir.iterdir() if path.is_file()]
    snapshot_paths = sorted(
        [path for path in paths if SNAPSHOT_FILENAME_RE.match(path.name)],
        key=lambda item: item.name,
    )
    if not snapshot_paths:
        raise PreviewRateError(
            "No completion snapshot files found. Expected names like 202603190700.xlsx."
        )
    return snapshot_paths


def run(workspace_dir: Path | None = None) -> Path:
    """Run the export workflow and return the output CSV path."""

    workspace = workspace_dir or get_workspace_dir()
    checkpoint_path = workspace / CHECKPOINT_FILENAME
    if not checkpoint_path.exists():
        raise PreviewRateError(f"Missing required checkpoint file: {checkpoint_path.name}")

    snapshot_paths = collect_snapshot_paths(workspace)
    checkpoint_videos = load_checkpoint_rows(checkpoint_path)
    if not checkpoint_videos:
        raise PreviewRateError("checkpoint.xlsx does not contain any active checkpoint rows.")

    snapshots = [load_completion_snapshot(path) for path in snapshot_paths]
    checkpoint_codes = sorted(checkpoint_videos)
    available_codes = {code for snapshot in snapshots for code in snapshot.video_columns}
    missing_codes = [code for code in checkpoint_codes if code not in available_codes]
    if missing_codes:
        raise PreviewRateError(
            "Checkpoint videos missing from all snapshots: " + ", ".join(missing_codes)
        )

    rows = build_student_unit_rows(checkpoint_videos, snapshots)
    output_path = workspace / OUTPUT_FILENAME
    write_csv(output_path, rows)

    unit_counts: dict[str, int] = {}
    for video in checkpoint_videos.values():
        unit_counts[video.unit_code] = unit_counts.get(video.unit_code, 0) + 1

    print(f"[ok] snapshots loaded: {len(snapshots)}")
    print(f"[ok] active checkpoint videos: {len(checkpoint_videos)}")
    print(f"[ok] output rows: {len(rows)}")
    for unit_code in sorted(unit_counts):
        print(f"[ok] {unit_code} eligible videos: {unit_counts[unit_code]}")
    print(f"[ok] wrote {output_path}")

    return output_path


def main() -> int:
    """CLI entrypoint."""

    if any(arg in {"-h", "--help"} for arg in sys.argv[1:]):
        print(HELP_TEXT)
        return 0

    try:
        run()
    except PreviewRateError as exc:
        print(f"[error] {exc}")
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
